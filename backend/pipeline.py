"""Wraps the ML pipeline as callable functions for the API."""

import logging
import sys
from concurrent.futures import ThreadPoolExecutor
from io import BytesIO
from pathlib import Path

logger = logging.getLogger(__name__)

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl.drawing.image import Image as XlImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from src.evaluation.metrics import evaluate_ranking
from src.models.common import MF_FEATURES
from src.models.logistic import logistic_model_fn, ridge_logistic_model_fn
from src.models.random_forest import detect_features, rf_model_fn
from src.models.vanilla_mf import mf_model_fn
from src.models.xgboost_model import xgb_model_fn

TOP_N = 30
MIN_TRAIN_PERIODS = 4

MODELS = [
    ("MagicFormula",   "A_MF_only",      MF_FEATURES, mf_model_fn),
    ("Logistic",       "B_MF_only",      MF_FEATURES, logistic_model_fn),
    ("Ridge",          "B_MF_only",      MF_FEATURES, ridge_logistic_model_fn),
    ("RandomForest",   "B_MF_only",      MF_FEATURES, rf_model_fn),
    ("XGBoost",        "B_MF_only",      MF_FEATURES, xgb_model_fn),
    ("Logistic",       "C_All_features", None,         logistic_model_fn),
    ("Ridge",          "C_All_features", None,         ridge_logistic_model_fn),
    ("RandomForest",   "C_All_features", None,         rf_model_fn),
    ("XGBoost",        "C_All_features", None,         xgb_model_fn),
]

METRIC_KEYS = [
    "accuracy", "precision", "recall", "f1", "roc_auc",
    "spearman_ic", "ndcg_at_k", "precision_at_k",
    "mean_annual_return", "std_annual_return",
    "cagr", "sharpe", "max_drawdown",
]

METRIC_LABELS = {
    "accuracy": "Accuracy", "precision": "Precision", "recall": "Recall",
    "f1": "F1 Score", "roc_auc": "ROC AUC", "spearman_ic": "Spearman IC",
    "ndcg_at_k": "NDCG@30", "precision_at_k": "Precision@30",
    "mean_annual_return": "Mean Ann. Return", "std_annual_return": "Std Ann. Return",
    "cagr": "CAGR", "sharpe": "Sharpe", "max_drawdown": "Max Drawdown",
}


def validate_csv(df):
    required = {"ticker", "quarter", "forward_return", "forward_return_rank",
                "earnings_yield", "return_on_capital"}
    missing = required - set(df.columns)
    if missing:
        raise ValueError(f"Missing columns: {', '.join(missing)}")
    features = detect_features(df)
    quarters = sorted(df["quarter"].unique())
    return {
        "rows": len(df), "quarters": len(quarters),
        "quarter_range": f"{quarters[0]} – {quarters[-1]}",
        "features": features,
        "columns": [
            {"name": c, "dtype": str(df[c].dtype),
             "sample": str(df[c].iloc[0]) if len(df) > 0 else ""}
            for c in df.columns
        ],
    }


def _run_fold(fn, train_df, test_df, feature_cols, is_mf):
    if is_mf:
        result = fn(train_df, test_df)
        result["label"] = (result["forward_return_rank"] >= 50).astype(int)
        model = None
    else:
        result, model = fn(train_df, test_df, feature_cols=feature_cols)
    m = evaluate_ranking(result, score_col="predicted_score",
                         return_col="forward_return_rank",
                         label_col="label", pred_label_col="predicted_label", k=TOP_N)
    # Store the single-quarter portfolio return for later aggregation.
    # Don't compute CAGR/Sharpe/Drawdown per-fold (meaningless for 1 quarter).
    if "forward_return" in result.columns:
        top = result.nlargest(min(TOP_N, len(result)), "predicted_score")
        avg_ret = top["forward_return"].mean()
        m["portfolio_return"] = float(avg_ret) if np.isfinite(avg_ret) else None
    return m, model


def _train_one_model(args):
    """Train a single model across all folds. Called from thread pool."""
    model_name, fs_name, features, fn, df, folds, q_idx, on_progress = args
    is_mf = model_name == "MagicFormula"
    fold_metrics, last_model = [], None

    for i, (train_qs, test_qs) in enumerate(folds):
        train = df.iloc[np.concatenate([q_idx[q] for q in train_qs])]
        test = df.iloc[q_idx[test_qs[0]]]
        if len(train) < 10 or len(test) < 5:
            continue
        if not is_mf and train["label"].nunique() < 2:
            continue
        try:
            m, model = _run_fold(fn, train, test, features, is_mf)
            m.update(quarter=test_qs[0], model=model_name, feature_set=fs_name)
            fold_metrics.append(m)
            if model:
                last_model = model
            if on_progress:
                on_progress(model_name, fs_name, i + 1, len(folds),
                            m.get("roc_auc", 0))
        except Exception as e:
            logger.warning("%s (%s) fold %d failed: %s", model_name, fs_name, i + 1, e)

    fi = None
    if fs_name.startswith("C_") and last_model:
        if hasattr(last_model, "feature_importances_"):
            imp = last_model.feature_importances_.tolist()
        elif hasattr(last_model, "coef_"):
            imp = np.abs(last_model.coef_[0]).tolist()
        else:
            imp = None
        if imp is not None:
            fi = {"features": list(features), "importances": imp}

    avg = None
    if fold_metrics:
        mdf = pd.DataFrame(fold_metrics)
        avg = {"model": model_name, "feature_set": fs_name}
        # Average classification/ranking metrics across folds.
        # Skip portfolio metrics — they are computed separately below.
        portfolio_keys = {"cagr", "sharpe", "max_drawdown",
                          "mean_annual_return", "std_annual_return"}
        for k in METRIC_KEYS:
            if k in portfolio_keys:
                continue
            if k in mdf.columns:
                vals = mdf[k].dropna()
                vals = vals[np.isfinite(vals)]
                avg[k] = round(float(vals.mean()), 6) if len(vals) > 0 else None
        avg["n_folds"] = len(fold_metrics)

        # Portfolio metrics.  forward_return is a 12-month return, so each
        # fold's portfolio_return is already annual-scale.  Consecutive folds
        # overlap (~9 months), so we report simple stats from all folds and
        # compute CAGR/Sharpe/Drawdown from non-overlapping Q1-only folds.
        all_rets = [m["portfolio_return"] for m in fold_metrics
                    if m.get("portfolio_return") is not None
                    and np.isfinite(m["portfolio_return"])]
        if len(all_rets) >= 2:
            r = np.array(all_rets)
            avg["mean_annual_return"] = round(float(np.mean(r)), 6)
            avg["std_annual_return"] = round(float(np.std(r, ddof=1)), 6)

        # Non-overlapping annual returns (Q1 folds only) for CAGR/Sharpe/Drawdown
        q1_rets = [m["portfolio_return"] for m in fold_metrics
                   if m.get("portfolio_return") is not None
                   and m.get("quarter", "").endswith("Q1")
                   and np.isfinite(m["portfolio_return"])]
        if len(q1_rets) >= 2:
            r_q1 = np.array(q1_rets)
            avg["cagr"] = round(float(np.prod(1 + r_q1) ** (1 / len(r_q1)) - 1), 6)
            q1_std = float(np.std(r_q1, ddof=1))
            avg["sharpe"] = round((float(np.mean(r_q1)) - 0.05) / q1_std, 6) if q1_std > 1e-10 else 0.0
            cum = np.cumprod(1 + r_q1)
            peak = np.maximum.accumulate(cum)
            avg["max_drawdown"] = round(float(np.min((cum - peak) / peak)), 6)
        else:
            avg["cagr"] = None
            avg["sharpe"] = None
            avg["max_drawdown"] = None

    return model_name, fold_metrics, avg, fi


def run_pipeline(df, on_progress=None, max_workers=3):
    """Run all models with parallel training. Returns (comparison, per_quarter, feature_importances)."""
    df = df.copy()
    df["label"] = (df["forward_return_rank"] >= 50).astype(int)
    quarters = sorted(df["quarter"].unique())
    folds = [(quarters[:i], [quarters[i]]) for i in range(MIN_TRAIN_PERIODS, len(quarters))]
    all_features = detect_features(df)
    q_idx = df.groupby("quarter").indices  # quarter -> row indices, computed once

    tasks = [(name, fs, cols or all_features, fn, df, folds, q_idx, on_progress)
             for name, fs, cols, fn in MODELS]

    all_pq, comparison, layer_c_models = [], [], {}

    with ThreadPoolExecutor(max_workers=max_workers) as pool:
        for model_name, fold_metrics, avg, fi in pool.map(_train_one_model, tasks):
            if avg:
                comparison.append(avg)
            if fi:
                layer_c_models[model_name] = fi
            all_pq.extend(fold_metrics)

    comp_df = pd.DataFrame(comparison)
    pq_cols = ["model", "feature_set", "quarter"] + METRIC_KEYS + ["portfolio_return"]
    pq_df = pd.DataFrame(all_pq)
    pq_df = pq_df[[c for c in pq_cols if c in pq_df.columns]]

    # equal-weight market benchmark (average return of all stocks each quarter)
    if "forward_return" in df.columns:
        bench = []
        for q in quarters[MIN_TRAIN_PERIODS:]:
            avg_ret = df.loc[df["quarter"] == q, "forward_return"].mean()
            if np.isfinite(avg_ret):
                bench.append({"model": "Market", "feature_set": "Benchmark",
                              "quarter": q, "portfolio_return": float(avg_ret)})
        if bench:
            pq_df = pd.concat([pq_df, pd.DataFrame(bench)], ignore_index=True)

    return comp_df, pq_df, layer_c_models


# ── excel export ──────────────────────────────────────────────────────────

HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
LAYER_FILLS = {
    "A": PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid"),
    "B": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
    "C": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
}
THIN_BORDER = Border(bottom=Side(style="thin", color="B0B0B0"),
                     right=Side(style="thin", color="B0B0B0"))


def _fig_to_bytes(fig):
    buf = BytesIO()
    fig.savefig(buf, format="png", dpi=150, bbox_inches="tight")
    plt.close(fig)
    buf.seek(0)
    return buf


def _chart_metric_bars(comp_df):
    show = ["roc_auc", "spearman_ic", "precision_at_k", "f1"]
    labels = [f"{r['model']}\n{r['feature_set']}" for _, r in comp_df.iterrows()]
    x = np.arange(len(labels))
    w = 0.18
    fig, ax = plt.subplots(figsize=(12, 5))
    for j, m in enumerate(show):
        ax.bar(x + j * w, comp_df[m].fillna(0).values, w,
               label=METRIC_LABELS[m], color=["#4472C4", "#ED7D31", "#70AD47", "#FFC000"][j])
    ax.set_xticks(x + w * 1.5)
    ax.set_xticklabels(labels, fontsize=8)
    ax.set_ylabel("Score")
    ax.set_title("Model Comparison: Key Metrics")
    ax.legend(loc="upper left", fontsize=8)
    fig.tight_layout()
    return _fig_to_bytes(fig)


def _chart_ic(pq_df):
    lc = pq_df[pq_df["feature_set"] == "C_All_features"]
    if lc.empty:
        return None
    models = lc["model"].unique()
    fig, axes = plt.subplots(len(models), 1, figsize=(12, 3.5 * len(models)), squeeze=False)
    for i, name in enumerate(models):
        ax = axes[i, 0]
        d = lc[lc["model"] == name].sort_values("quarter")
        ics = d["spearman_ic"].values
        ax.bar(range(len(ics)), ics, color=["#4472C4" if v >= 0 else "#C44444" for v in ics], alpha=0.8)
        ax.axhline(np.nanmean(ics), color="red", ls="--", label=f"Mean IC = {np.nanmean(ics):.4f}")
        ax.axhline(0, color="black", lw=0.5)
        ticks = list(range(0, len(d), max(1, len(d) // 8)))
        ax.set_xticks(ticks)
        ax.set_xticklabels(d["quarter"].iloc[ticks], rotation=45, ha="right", fontsize=7)
        ax.set_title(f"{name} (All Features): IC per Quarter", fontsize=10)
        ax.set_ylabel("IC")
        ax.legend(fontsize=8)
    fig.tight_layout()
    return _fig_to_bytes(fig)


def _chart_fi(fi_dict):
    items = list(fi_dict.items())
    if not items:
        return None
    fig, axes = plt.subplots(1, len(items), figsize=(6 * len(items), max(4, len(items[0][1]["features"]) * 0.35)))
    if len(items) == 1:
        axes = [axes]
    for ax, (name, data) in zip(axes, items):
        imp = np.array(data["importances"])
        names = data["features"]
        idx = np.argsort(imp)
        ax.barh(range(len(names)), imp[idx], color="#4472C4")
        ax.set_yticks(range(len(names)))
        ax.set_yticklabels([names[i] for i in idx], fontsize=8)
        ax.set_xlabel("Importance (Gini)")
        ax.set_title(f"{name}: Feature Importance")
    fig.tight_layout()
    return _fig_to_bytes(fig)


def build_excel(comp_df, pq_df, fi_dict):
    """Returns BytesIO of the formatted xlsx."""
    buf = BytesIO()
    display_cols = {"model": "Model", "feature_set": "Layer", "n_folds": "Folds"}
    display_cols.update(METRIC_LABELS)

    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        comp_df.rename(columns=display_cols).to_excel(w, sheet_name="Comparison", index=False)
        pq_df.rename(columns={**display_cols, "quarter": "Quarter"}).to_excel(
            w, sheet_name="Per-Quarter", index=False)
        wb = w.book

        # style comparison
        ws = wb["Comparison"]
        nr, nc = len(comp_df), len(comp_df.columns)
        for col in range(1, nc + 1):
            c = ws.cell(row=1, column=col)
            c.fill, c.font, c.alignment = HEADER_FILL, HEADER_FONT, Alignment(horizontal="center")
        for row in range(2, nr + 2):
            layer = str(ws.cell(row=row, column=2).value or "")[0]
            fill = LAYER_FILLS.get(layer)
            for col in range(1, nc + 1):
                c = ws.cell(row=row, column=col)
                c.border = THIN_BORDER
                c.alignment = Alignment(horizontal="center")
                if fill:
                    c.fill = fill
                if col >= 3 and c.value is not None:
                    c.number_format = "0.0000"
        for col in range(1, nc + 1):
            mlen = max(len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, nr + 2))
            ws.column_dimensions[get_column_letter(col)].width = min(mlen + 3, 18)

        # embed charts
        img = _chart_metric_bars(comp_df)
        ws.add_image(XlImage(img), f"A{nr + 3}")

        ic_img = _chart_ic(pq_df)
        if ic_img:
            wb.create_sheet("IC Over Time")
            ws2 = wb["IC Over Time"]
            ws2["A1"] = "Spearman IC per quarter — Layer C models"
            ws2["A1"].font = Font(bold=True, size=11)
            ws2.add_image(XlImage(ic_img), "A3")

        fi_img = _chart_fi(fi_dict)
        if fi_img:
            wb.create_sheet("Feature Importance")
            ws3 = wb["Feature Importance"]
            ws3["A1"] = "Feature importance (Gini) — last walk-forward fold"
            ws3["A1"].font = Font(bold=True, size=11)
            ws3.add_image(XlImage(fi_img), "A3")

    buf.seek(0)
    return buf
